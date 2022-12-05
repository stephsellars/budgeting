"""The idea of this project is to create a program that will take
four inputs: month, day, type, and amount, and will enter the amount
in the correct cell according to the other three variables."""

"""Created as a final project for CSCS-1301"""

import csv
import openpyxl
from openpyxl import load_workbook

filename = "2023 Finances.xlsx"
wb = openpyxl.load_workbook(filename, read_only=False, keep_vba=True)

ws = wb.active

jan = wb["JAN"]
feb = wb["FEB"]
mar = wb["MAR"]
apr = wb["APR"]
may = wb["MAY"]
jun = wb["JUN"]
jul = wb["JUL"]
aug = wb["AUG"]
sep = wb["SEP"]
oct = wb["OCT"]
nov = wb["NOV"]
dec = wb["DEC"]
transactions = ["Paycheck 1", "Paycheck 2", "Bills", "Entertainment", "Food & Drink", "Groceries", "Health & Fitness"]

def greeting():
    answer = ""
    execute = True
    while execute:
        run_file = True
        while run_file:
            answer = input("Would you like to update your 2023 Finances?: (Y/N) \n")
            if answer.lower() == "yes" or answer.lower() == "y":
                print("Let's do it!\n")
                months()
            elif answer.lower() == "no" or answer.lower() == "n":
                print("Then why did you run the script?? okay.. goodbye")
                quit()
            else:
                print("That was a yes or no question! Please try again \n")    
    print("goodbye!")

def months():
    answer = input("What is the current month?: \n")
    if answer.lower() == "january":
        wb.active = jan
        days(jan)
    elif answer.lower() == "february":
        wb.active = feb
        days(feb)
    elif answer.lower() == "march":
        wb.active = mar
        days(mar)
    elif answer.lower() == "april":
        wb.active = apr
        days(apr)
    elif answer.lower() == "may":
        wb.active = may
        days(may)
    elif answer.lower() == "june":
        wb.active = jun
        days(jun)
    elif answer.lower() == "july":
        wb.active = jul
        days(jul)
    elif answer.lower() == "august":
        wb.active = aug
        days(aug)
    elif answer.lower() == "september":
        wb.active = sep
        days(sep)
    elif answer.lower() == "october":
        wb.active = oct
        days(oct)
    elif answer.lower() == "november":
        wb.active = nov
        days(nov)
    elif answer.lower() == "december":
        wb.active = dec
        days(dec)
    return answer
    quit()


def days(month):
    n = input("What day of the month is it?: \n")
    if n is int:
        mycell()
    types()
    return n
    
def types():
    d = input(f"What type of transaction is this?: {transactions}\n")
    confirm = input(f"You chose {d}, is this correct?\n")
    if confirm == "yes".lower():
            amount()
    else:
        types()
    return d
    
def amount():
    money = input("How much was the transaction?:\n")
    confirm = input(f"Is this the right amount? ${money}\n")     
    if confirm == "yes".lower() or confirm == "y".lower():
        goodbye()
    else:
        amount()
    return money
    
def goodbye():
    print("Thank you for using this script!! \nYour excel sheet has been updated")
    wb.save(filename)
    wb.close()
    quit()

def mycell():
    sheetname = wb.get_sheet_by_name(months())
    sheetname.cell(row=days(), column=types()).value = amount()



print(greeting())
